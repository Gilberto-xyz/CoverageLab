import csv
import io
import os
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NORMALIZED_INPUT_COLUMNS = [
    "date_raw",
    "sell_out",
    "penetracion",
    "compra_media",
    "compra_por_ocasion",
    "frecuencia",
    "buyers",
    "sell_in",
]

COUNTRY_MAP = {
    "10": "LatAm",
    "54": "Argentina",
    "91": "Bolivia",
    "55": "Brasil",
    "12": "CAM",
    "56": "Chile",
    "57": "Colombia",
    "93": "Ecuador",
    "52": "Mexico",
    "51": "Peru",
    "69": "Republica Dominicana",
    "62": "Guatemala",
    "63": "El Salvador",
    "64": "Honduras",
    "65": "Nicaragua",
    "66": "Costa Rica",
    "67": "Panamá",
}

EXCEL_TEMP_FILENAME = "file_temp_coverage.xlsx"
DEFAULT_SAMPLE_SIZE = 1200
DEFAULT_COVERAGE_LABEL = "Cobertura Absoluta"

POP_COVERAGE_MAP = {
    "Argentina": "90%",
    "Bolivia": "60%",
    "Brasil": "82%",
    "Chile": "78%",
    "Colombia": "65%",
    "Ecuador": "61%",
    "Mexico": "64%",
    "Peru": "66%",
    "CAM": "74%",
    "Costa Rica": "94%",
    "El Salvador": "86%",
    "Guatemala": "69%",
    "Honduras": "65%",
    "Nicaragua": "57%",
    "Panama": "92%",
    "Republica Dominicana": "63%",
}

BRASIL_BENCHMARK_BY_PENETRATION = [
    ("0% - 5%", "40%"),
    ("6% - 10%", "53%"),
    ("11% - 30%", "54%"),
    ("31% - 70%", "59%"),
]


def _load_categories_map_from_coverage_studio():
    studio_path = Path(__file__).with_name("coverage_studio.py")
    if not studio_path.exists():
        return {}

    try:
        text = studio_path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return {}

    match = re.search(r'CATEGORIES_CSV_DATA\s*=\s*"""(.*?)"""', text, flags=re.DOTALL)
    if not match:
        return {}

    csv_blob = match.group(1).strip()
    categories = {}
    try:
        reader = csv.DictReader(io.StringIO(csv_blob))
        for row in reader:
            code = str(row.get("cod", "")).strip()
            cat = str(row.get("cat", "")).strip()
            if code:
                categories[code] = cat
    except Exception:
        return {}
    return categories


CATEGORY_NAME_MAP = _load_categories_map_from_coverage_studio()


class Colors:
    HEADER = "\033[95m"
    OKBLUE = "\033[94m"
    OKCYAN = "\033[96m"
    OKGREEN = "\033[92m"
    WARNING = "\033[93m"
    FAIL = "\033[91m"
    ENDC = "\033[0m"
    BOLD = "\033[1m"


def _safe_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if text in {"", "-"}:
        return None
    text = text.replace("%", "").replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def _select_from_list(title, options):
    if not options:
        raise ValueError("No hay opciones disponibles para seleccionar.")

    print(f"\n{Colors.OKCYAN}{title}{Colors.ENDC}")
    for idx, option in enumerate(options, 1):
        print(f"{Colors.OKGREEN}[{idx}]{Colors.ENDC} {option}")

    while True:
        choice = input(f"\n{Colors.OKBLUE}Selecciona una opcion: {Colors.ENDC}").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(options):
            idx = int(choice) - 1
            return idx, options[idx]
        print(f"{Colors.FAIL}Entrada invalida. Intenta de nuevo.{Colors.ENDC}")


def _quick_file_metadata(filename):
    parts = Path(filename).stem.split("_")
    if len(parts) < 2:
        return ""
    country = COUNTRY_MAP.get(parts[0], "Desconocido")
    return f"{country} | Categoria: {parts[1]}"


def _select_excel_files():
    files = sorted(
        [
            f
            for f in os.listdir(".")
            if f.lower().endswith(".xlsx") and not f.startswith("~$") and f != EXCEL_TEMP_FILENAME
        ]
    )
    if not files:
        raise FileNotFoundError("No se encontraron archivos .xlsx para procesar.")

    print(f"\n{Colors.OKCYAN}Archivos Excel (.xlsx) disponibles:{Colors.ENDC}")
    for idx, filename in enumerate(files, 1):
        meta = _quick_file_metadata(filename)
        if meta:
            print(f"{Colors.OKGREEN}[{idx}]{Colors.ENDC} {filename} {Colors.WARNING}| {meta}{Colors.ENDC}")
        else:
            print(f"{Colors.OKGREEN}[{idx}]{Colors.ENDC} {filename}")

    while True:
        choice = input(
            f"\n{Colors.OKBLUE}Selecciona archivo(s) (1-{len(files)}), separados por coma, o escribe 'all': {Colors.ENDC}"
        ).strip().lower()
        if choice in {"all", "todos", "*"}:
            return files
        try:
            selected_indices = [int(x.strip()) for x in choice.split(",") if x.strip()]
        except ValueError:
            print(f"{Colors.FAIL}Entrada invalida. Ingrese números separados por coma o 'all'.{Colors.ENDC}")
            continue
        if not selected_indices:
            print(f"{Colors.FAIL}Debe seleccionar al menos un archivo.{Colors.ENDC}")
            continue
        if not all(1 <= idx <= len(files) for idx in selected_indices):
            print(f"{Colors.FAIL}Uno o más índices están fuera de rango.{Colors.ENDC}")
            continue
        return [files[idx - 1] for idx in selected_indices]


def _select_country():
    countries = sorted(COUNTRY_MAP.values())
    _, selected = _select_from_list("Paises disponibles:", countries)
    return selected


def _select_criterio():
    _, selected = _select_from_list("Criterio Scorecard Unilever:", ["Si", "No"])
    return selected


def _select_output_name(default_name):
    print(f"\n{Colors.OKCYAN}Nombre de archivo de salida (automatico):{Colors.ENDC} {Colors.OKGREEN}{default_name}{Colors.ENDC}")
    return default_name


def _parse_start_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value

    text = str(value).strip()
    for fmt in ("%b-%y  ", "%b-%y", "%m-%y", "%Y-%m-%d"):
        try:
            return pd.Timestamp(datetime.strptime(text, fmt))
        except ValueError:
            continue

    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed


def _build_month_labels(first_value, n_rows):
    start = _parse_start_date(first_value)
    if start is None:
        return [f"M{idx+1:02d}" for idx in range(n_rows)]
    return [(start + pd.DateOffset(months=idx)).strftime("%m-%y") for idx in range(n_rows)]


def _normalize_text(value):
    text = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip()


def _country_from_source_file(source_file):
    stem = Path(source_file).stem
    parts = re.split(r"[_\-]+", stem)
    if not parts:
        return None

    # Intento 1: prefijo por código de país (ej. 55_xxxx).
    by_code = COUNTRY_MAP.get(parts[0])
    if by_code:
        return by_code

    # Intento 2: país explícito en el nombre (ej. Brasil_xxxx o Brazil-xxxx).
    normalized_stem = _normalize_text(stem)
    country_aliases = {
        "brasil": "Brasil",
        "brazil": "Brasil",
    }
    for country in COUNTRY_MAP.values():
        country_aliases[_normalize_text(country)] = country

    for alias, canonical in country_aliases.items():
        if alias and alias in normalized_stem:
            return canonical
    return None


def _select_country_for_source(source_file):
    inferred_country = _country_from_source_file(source_file)
    if inferred_country:
        print(f"{Colors.OKGREEN}Pais detectado por nombre de archivo:{Colors.ENDC} {inferred_country}")
        return inferred_country
    print(f"{Colors.WARNING}No se pudo detectar país desde el nombre '{source_file}'. Seleccione manualmente.{Colors.ENDC}")
    return _select_country()


def _infer_sample_size(brand_df):
    buyers = pd.to_numeric(brand_df.get("buyers"), errors="coerce")
    penet = pd.to_numeric(brand_df.get("penetracion"), errors="coerce")
    valid = buyers.notna() & penet.notna() & (buyers > 0) & (penet > 0)
    if not valid.any():
        return None

    inferred = buyers[valid] / (penet[valid] / 100.0)
    inferred = inferred.replace([float("inf"), float("-inf")], pd.NA).dropna()
    if inferred.empty:
        return None

    # Tomamos la mediana para robustez frente a outliers.
    value = float(inferred.median())
    if value <= 0:
        return None
    return int(round(value))


def _print_brazil_benchmark_notification():
    message = (
        "Notificacion: Se detecto Brasil. "
        "Se utilizaran los valores del Benchmark de Cobertura por nivel de penetracion."
    )
    print(f"\n{Colors.WARNING}{message}{Colors.ENDC}")

    headers = ("Rango de Penetracion", "Cobertura Benchmark")
    col1_width = max(len(headers[0]), max(len(row[0]) for row in BRASIL_BENCHMARK_BY_PENETRATION))
    col2_width = max(len(headers[1]), max(len(row[1]) for row in BRASIL_BENCHMARK_BY_PENETRATION))
    separator = "-" * (col1_width + col2_width + 3)

    print(f"{Colors.OKCYAN}{separator}{Colors.ENDC}")
    print(f"{Colors.OKCYAN}{headers[0].ljust(col1_width)} | {headers[1].ljust(col2_width)}{Colors.ENDC}")
    print(f"{Colors.OKCYAN}{separator}{Colors.ENDC}")
    for penetration_range, coverage in BRASIL_BENCHMARK_BY_PENETRATION:
        print(f"{penetration_range.ljust(col1_width)} | {coverage.ljust(col2_width)}")
    print(f"{Colors.OKCYAN}{separator}{Colors.ENDC}")
    print("Nota: Fuera de esos rangos, se aplica 82% como cobertura base para Brasil.")


def _get_population_coverage(pais, penet1):
    if _normalize_text(pais) == "brasil":
        penet = _safe_float(penet1)
        if penet is None:
            return 82
        if 0 <= penet <= 5:
            return 40
        if 5 < penet <= 10:
            return 53
        if 10 < penet <= 30:
            return 54
        if 30 < penet <= 70:
            return 59
        return 82

    target = _normalize_text(pais)
    for country_name, value in POP_COVERAGE_MAP.items():
        if _normalize_text(country_name) == target:
            return float(str(value).replace("%", "").strip())
    return 100.0


def _extract_table_block(raw_df, brand):
    start_idx = 0
    try:
        first_col = raw_df.iloc[:, 0].astype(str)
        table_mask = first_col.str.contains(r"\btable\b", flags=re.IGNORECASE, na=False)
        if table_mask.any():
            start_idx = int(table_mask[table_mask].index[0])
    except Exception:
        start_idx = 0

    df = raw_df.iloc[start_idx:, :].reset_index(drop=True)
    rows, cols = df.shape

    if rows < 2 or cols < 8:
        raise ValueError(
            f"La hoja '{brand}' tiene estructura invalida ({rows} filas, {cols} columnas). "
            "Se requieren al menos 2 filas y 8 columnas."
        )

    try:
        col8 = df.iloc[1:, 7]
        col8_empty = col8.isna().all() or col8.astype(str).str.strip().eq("").all()
    except Exception:
        col8_empty = True

    if col8_empty:
        raise ValueError(
            f"La hoja '{brand}' no tiene datos en la columna 8 (Sell-in) debajo del encabezado."
        )

    return df


def _normalize_brand_sheet(raw_df, brand, category_name):
    raw_table = _extract_table_block(raw_df, brand)
    df = raw_table.iloc[1:, :8].copy()
    df.columns = NORMALIZED_INPUT_COLUMNS
    df = df.dropna(how="all").reset_index(drop=True)

    if df.empty:
        raise ValueError(f"La hoja '{brand}' no tiene datos.")

    parsed_dates = pd.to_datetime(df["date_raw"].apply(_parse_start_date), errors="coerce")
    valid_dates = parsed_dates.notna().sum()
    if valid_dates == len(df):
        df["date"] = parsed_dates.dt.strftime("%m-%y")
    elif valid_dates == 1 and pd.notna(parsed_dates.iloc[0]):
        df["date"] = _build_month_labels(parsed_dates.iloc[0], len(df))
    elif valid_dates > 0:
        valid_mask = parsed_dates.notna()
        df = df.loc[valid_mask].reset_index(drop=True)
        parsed_dates = parsed_dates.loc[valid_mask].reset_index(drop=True)
        df["date"] = parsed_dates.dt.strftime("%m-%y")
    else:
        df["date"] = _build_month_labels(df["date_raw"].iloc[0], len(df))

    for col in ["sell_out", "penetracion", "compra_media", "compra_por_ocasion", "frecuencia", "buyers", "sell_in"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["Marca"] = brand
    df["Categoria"] = category_name

    return df[
        [
            "date",
            "sell_out",
            "penetracion",
            "compra_media",
            "compra_por_ocasion",
            "frecuencia",
            "buyers",
            "sell_in",
            "Marca",
            "Categoria",
        ]
    ]


def _calc_pipeline(df_brand):
    pipe = pd.DataFrame(index=df_brand.index)
    pipe["Sell-In MAT"] = df_brand["sell_in"].rolling(window=12).sum().round(1)
    pipe["Sell-Out MAT"] = df_brand["sell_out"].rolling(window=12).sum().round(1)
    pipe = pipe.dropna().copy()

    for i in range(7):
        pipe[f"Pipeline {i}"] = (pipe["Sell-Out MAT"] / pipe["Sell-In MAT"].shift(i) * 100).round(2)

    pipe = pipe.join(df_brand[["date", "penetracion", "buyers", "Marca", "Categoria"]], how="left")
    pipe = pipe.dropna(subset=["Pipeline 0"]).reset_index(drop=True)

    return pipe[
        [
            "date",
            "Marca",
            "Categoria",
            "penetracion",
            "buyers",
            "Pipeline 0",
            "Pipeline 1",
            "Pipeline 2",
            "Pipeline 3",
            "Pipeline 4",
            "Pipeline 5",
            "Pipeline 6",
        ]
    ]


def _skipped_sheet(sheet_name, stage, exc):
    error = (re.sub(r"\s+", " ", str(exc)).strip() or type(exc).__name__).rstrip(".")
    return {"sheet": str(sheet_name), "stage": stage, "error": error}


def _notify_skipped_sheet(skipped):
    print(
        f"{Colors.WARNING}Aviso: se omitió la hoja '{skipped['sheet']}' "
        f"durante {skipped['stage']}: {skipped['error']}. "
        f"El proceso continuará con las demás hojas.{Colors.ENDC}"
    )


def _print_skipped_sheets_summary(source_file, skipped_sheets):
    if not skipped_sheets:
        return

    print(f"\n{Colors.WARNING}Resumen de hojas omitidas en '{source_file}':{Colors.ENDC}")
    for skipped in skipped_sheets:
        print(f"- {skipped['sheet']} ({skipped['stage']}): {skipped['error']}")


def _load_source_data(source_file):
    category_name = Path(source_file).stem
    pipeline_by_brand = {}
    all_pipelines = []
    valid_sheet_names = []
    skipped_sheets = []
    with pd.ExcelFile(source_file) as xls:
        for sheet_name in xls.sheet_names:
            try:
                brand_name = _clean_brand_name_from_sheet(sheet_name)
                raw_df = pd.read_excel(xls, header=None, sheet_name=sheet_name)
                normalized = _normalize_brand_sheet(raw_df, brand_name, category_name)
                pipe = _calc_pipeline(normalized)
                if len(pipe) < 13:
                    raise ValueError(f"La marca '{brand_name}' no tiene suficientes periodos para scorecard (mínimo 13).")
                pipeline_by_brand[sheet_name] = pipe
                all_pipelines.append(pipe)
                valid_sheet_names.append(sheet_name)
            except Exception as exc:
                skipped = _skipped_sheet(sheet_name, "la carga", exc)
                skipped_sheets.append(skipped)
                _notify_skipped_sheet(skipped)

    total_pipeline = pd.concat(all_pipelines, ignore_index=True) if all_pipelines else pd.DataFrame()
    return category_name, valid_sheet_names, total_pipeline, pipeline_by_brand, skipped_sheets


def _extract_preassigned_pipeline(sheet_name):
    match = re.match(r"(?i)^p([0-6])_", str(sheet_name or "").strip())
    return int(match.group(1)) if match else None


def _pipelines_to_run_for_sheet(sheet_name):
    preassigned = _extract_preassigned_pipeline(sheet_name)
    return [preassigned] if preassigned is not None else list(range(7))


def _clean_brand_name_from_sheet(sheet_name):
    cleaned = re.sub(r"(?i)^p[0-6]_", "", str(sheet_name or "")).strip()
    return cleaned or str(sheet_name or "N/D")


def _short_category_name(category_name):
    try:
        parts = re.split(r"\s*[-‑–—−‒]\s*", str(category_name or ""), maxsplit=1)
        short_name = (parts[0] if parts else str(category_name or "")).strip()
    except Exception:
        short_name = str(category_name or "").strip()
    return short_name or str(category_name or "").strip()


def _parse_input_metadata(source_file):
    parts = Path(source_file).stem.split("_")
    if len(parts) < 3:
        raise ValueError("El nombre de archivo no contiene suficientes partes (país_categoria_fabricante)")

    country_code = parts[0]
    category_code = parts[1]
    manufacturer = parts[2]

    country_name = COUNTRY_MAP.get(country_code, f"Pais-{country_code}")
    category_name = CATEGORY_NAME_MAP.get(category_code, category_code)
    category_short = _short_category_name(category_name)

    return {
        "country_name": country_name,
        "category_code": category_code,
        "category_name": category_name,
        "category_short": category_short,
        "manufacturer": manufacturer,
    }


def _infer_reference_month_year(sheet_names, pipeline_by_brand):
    for sheet_name in reversed(sheet_names):
        df_brand = pipeline_by_brand.get(sheet_name)
        if df_brand is None or df_brand.empty:
            continue
        raw_last = df_brand["date"].iloc[-1]
        parsed = _parse_start_date(raw_last)
        if parsed is not None and not pd.isna(parsed):
            return parsed.strftime("%m-%y")
        raw_text = str(raw_last).strip()
        if re.match(r"^\d{2}-\d{2}$", raw_text):
            return raw_text
    return datetime.now().strftime("%m-%y")


def _resolve_output_target(source_file, sheet_names, pipeline_by_brand, criterio):
    metadata = _parse_input_metadata(source_file)
    ref_month_year = _infer_reference_month_year(sheet_names, pipeline_by_brand)
    base_name = (
        f"{metadata['country_name']}-{metadata['category_short']}-"
        f"{metadata['manufacturer']}-{ref_month_year}_{DEFAULT_COVERAGE_LABEL}"
    )
    output_dir = Path(__file__).resolve().parent / base_name
    criterio_suffix = "unilever" if _normalize_text(criterio) in {"si", "unilever"} else "personalizado"
    default_output_name = f"Scorecard_{base_name}_{criterio_suffix}.xlsx"
    return output_dir, default_output_name


def _compute_scorecard(brand_df, brand, pipeline, pais, criterio):
    fecha1 = str(brand_df["date"].iloc[-1])
    fecha2 = str(brand_df["date"].iloc[-13])

    pipe1 = float(brand_df[f"Pipeline {pipeline}"].iloc[-1])
    pipe2 = float(brand_df[f"Pipeline {pipeline}"].iloc[-13])
    estab = abs(round(pipe1 - pipe2, 1))

    penet1 = float(brand_df["penetracion"].iloc[-12:].mean())
    penet2 = float(brand_df["penetracion"].iloc[-24:-12].mean()) if len(brand_df) >= 24 else penet1

    muestra_total = _infer_sample_size(brand_df) or DEFAULT_SAMPLE_SIZE

    error_standar1 = ((((penet1 / 100) * (1 - (penet1 / 100))) / muestra_total) ** 0.5) * 1.96 if muestra_total > 0 else 0
    error_standar2 = ((((penet2 / 100) * (1 - (penet2 / 100))) / muestra_total) ** 0.5) * 1.96 if muestra_total > 0 else 0

    error_relativo1 = (error_standar1 / penet1) * 100 if penet1 not in (0, None) else 0
    error_relativo2 = (error_standar2 / penet2) * 100 if penet2 not in (0, None) else 0
    _ = error_relativo2  # Kept for parity with original formulas.

    cobertura_poblacional = _get_population_coverage(pais, penet1)

    if criterio == "No":
        limit_inferior_verde = int(round((cobertura_poblacional - (cobertura_poblacional * error_relativo1)), 2))
        limit_superior_verde = int(round((100 + (cobertura_poblacional * error_relativo1)), 2))
        diff = cobertura_poblacional - limit_inferior_verde
        limit_inferior_rojo = limit_inferior_verde - diff
        limit_superior_rojo = limit_superior_verde + diff
        limite_inf_estabilidad = int(round((pipe1 + (pipe1 * error_relativo1)), 0) - round((pipe1 - (pipe1 * error_relativo1)), 0)) + 1
    else:
        limit_inferior_verde = 80
        limit_superior_verde = 110
        limit_inferior_rojo = 50
        limit_superior_rojo = 120
        limite_inf_estabilidad = 5

    cob_excelente = f">= {limit_inferior_verde}% - <{limit_superior_verde}%"
    cob_bueno = f"{limit_inferior_rojo} - {limit_inferior_verde}% ; {limit_superior_verde} - {limit_superior_rojo}%"
    cob_alerta = f">= {limit_inferior_rojo}% - <{limit_superior_rojo}%"

    est_excelente = f"<= {limite_inf_estabilidad}"
    est_bueno = f"{limite_inf_estabilidad} - {limite_inf_estabilidad * 2}"
    est_alerta = f"> {limite_inf_estabilidad * 2 + 1}"

    criterio_nombre = "Bom" if pais == "Brasil" else "Bueno"
    criterio_estab = "Estabilidade" if pais == "Brasil" else "Estabilidad"

    scorecard_df = pd.DataFrame(
        {
            "Score Card": ["Cobertura", criterio_estab],
            fecha2: [str(round(pipe2, 1)), "-"],
            fecha1: [str(round(pipe1, 1)), str(estab)],
            "Excelente": [cob_excelente, est_excelente],
            criterio_nombre: [cob_bueno, est_bueno],
            "Alerta": [cob_alerta, est_alerta],
        }
    ).set_index("Score Card")

    return {
        "marca": brand,
        "pipeline": pipeline,
        "scorecard": scorecard_df,
        "penet_12m": float(round(penet1, 2)),
        "cobertura_benchmark": float(round(cobertura_poblacional, 2)),
        "lim_inf_verde": float(limit_inferior_verde),
        "lim_sup_verde": float(limit_superior_verde),
        "lim_inf_rojo": float(limit_inferior_rojo),
        "lim_sup_rojo": float(limit_superior_rojo),
        "lim_estab": float(limite_inf_estabilidad),
        "muestra_total": int(muestra_total),
    }


def _build_scorecards(pais, criterio, sheet_names, pipeline_by_brand):
    entries = []
    skipped_sheets = []
    total = sum(len(_pipelines_to_run_for_sheet(sheet_name)) for sheet_name in sheet_names)

    print(f"\n{Colors.OKCYAN}Calculando scorecards...{Colors.ENDC}")
    start = time.perf_counter()
    progress = 0

    for sheet_name in sheet_names:
        brand_name = _clean_brand_name_from_sheet(sheet_name)
        pipelines = _pipelines_to_run_for_sheet(sheet_name)
        sheet_entries = []
        attempted = 0

        try:
            brand_df = pipeline_by_brand[sheet_name]
            for attempted, pipeline in enumerate(pipelines, 1):
                progress += 1
                print(f"{Colors.OKBLUE}[{progress}/{total}]{Colors.ENDC} Marca: {brand_name} | Pipeline: {pipeline}")
                entry = _compute_scorecard(brand_df, brand_name, pipeline, pais, criterio)
                sheet_entries.append(entry)
        except Exception as exc:
            progress += len(pipelines) - attempted
            skipped = _skipped_sheet(sheet_name, "el cálculo", exc)
            skipped_sheets.append(skipped)
            _notify_skipped_sheet(skipped)
            continue

        entries.extend(sheet_entries)

    elapsed = time.perf_counter() - start
    print(f"{Colors.OKGREEN}Scorecards calculados en {elapsed:.1f}s.{Colors.ENDC}")
    return entries, skipped_sheets


def export_scorecards_single_sheet(scorecards, output_dir, output_file, sheet_name="Scorecards"):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    fills = {
        "header": PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid"),
        "row_label": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "green": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        "yellow": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "red": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    }

    header_font = Font(color="FFFFFF", bold=True)
    normal_font = Font(color="000000")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for entry in scorecards:
        scorecard_df = entry["scorecard"].copy()
        scorecard_df.insert(0, "Score Card", scorecard_df.index)

        headers = ["Marca", "Pipeline"] + list(scorecard_df.columns) + ["", "Penetracion Prom 12M", "Cobertura Benchmark"]
        sheet.append(headers)
        header_row = sheet.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=header_row, column=col_idx)
            cell.fill = fills["header"]
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

        start_row = sheet.max_row + 1
        for values in scorecard_df.itertuples(index=False, name=None):
            penet_value = float(entry.get("penet_12m", 0.0))
            cobertura_benchmark = float(entry.get("cobertura_benchmark", 0.0))
            sheet.append([entry["marca"], entry["pipeline"]] + list(values) + [None, penet_value, cobertura_benchmark])
        end_row = sheet.max_row

        if end_row > start_row:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            sheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            sheet.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
            sheet.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)

        coverage_row = None
        stab_row = None
        for row_idx in range(start_row, end_row + 1):
            label = str(sheet.cell(row=row_idx, column=3).value).lower()
            if "cobertura" in label:
                coverage_row = row_idx
            elif "estab" in label:
                stab_row = row_idx

        if coverage_row is not None:
            for col_idx in [4, 5]:
                value = _safe_float(sheet.cell(row=coverage_row, column=col_idx).value)
                if value is None:
                    continue
                if entry["lim_inf_verde"] <= value <= entry["lim_sup_verde"]:
                    sheet.cell(row=coverage_row, column=col_idx).fill = fills["green"]
                elif value < entry["lim_inf_rojo"] or value > entry["lim_sup_rojo"]:
                    sheet.cell(row=coverage_row, column=col_idx).fill = fills["red"]
                else:
                    sheet.cell(row=coverage_row, column=col_idx).fill = fills["yellow"]

        if stab_row is not None:
            value = _safe_float(sheet.cell(row=stab_row, column=5).value)
            if value is not None:
                if value <= entry["lim_estab"]:
                    sheet.cell(row=stab_row, column=5).fill = fills["green"]
                elif value > (2 * entry["lim_estab"]):
                    sheet.cell(row=stab_row, column=5).fill = fills["red"]
                else:
                    sheet.cell(row=stab_row, column=5).fill = fills["yellow"]

        for row_idx in range(start_row, end_row + 1):
            sheet.cell(row=row_idx, column=3).fill = fills["row_label"]
            sheet.cell(row=row_idx, column=6).fill = fills["green"]
            sheet.cell(row=row_idx, column=7).fill = fills["yellow"]
            sheet.cell(row=row_idx, column=8).fill = fills["red"]
            sheet.cell(row=row_idx, column=10).number_format = '0.0"%"'
            sheet.cell(row=row_idx, column=11).number_format = '0.0"%"'

            for col_idx in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.font = normal_font
                cell.alignment = align_center
                cell.border = thin_border

        sheet.append([None] * len(headers))

    for col_idx in range(1, sheet.max_column + 1):
        max_len = 0
        for row_idx in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 38)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_file
    wb.save(output_path)
    print(f"\n{Colors.OKGREEN}Scorecards exportados en una sola hoja:{Colors.ENDC} {output_path}")
    return output_path


def main():
    print(f"{Colors.HEADER}{Colors.BOLD}=== Exportador Interactivo de Scorecards (Autocontenido) ==={Colors.ENDC}")
    global_start = time.perf_counter()

    auto_file = os.environ.get("AUTO_FILE")
    if auto_file:
        source_files = [Path(auto_file).name if Path(auto_file).exists() else auto_file]
        print(f"\n{Colors.OKCYAN}AUTO_FILE detectado:{Colors.ENDC} {source_files[0]}")
    else:
        source_files = _select_excel_files()

    criterio = _select_criterio()
    total_files = len(source_files)

    for idx, source_file in enumerate(source_files, 1):
        print(f"\n{Colors.OKCYAN}=== Archivo {idx}/{total_files}: {source_file} ==={Colors.ENDC}")
        try:
            category_name, sheet_names, _, pipeline_by_brand, skipped_sheets = _load_source_data(source_file)
            if not sheet_names:
                print(f"{Colors.FAIL}No se encontraron hojas válidas en '{source_file}'. Se omitirá el archivo.{Colors.ENDC}")
                _print_skipped_sheets_summary(source_file, skipped_sheets)
                continue

            pais = _select_country_for_source(source_file)
            if _normalize_text(pais) == "brasil":
                _print_brazil_benchmark_notification()
            output_dir, default_output_name = _resolve_output_target(
                source_file=source_file,
                sheet_names=sheet_names,
                pipeline_by_brand=pipeline_by_brand,
                criterio=criterio,
            )

            output_name = _select_output_name(default_output_name)

            scorecards, calculation_skips = _build_scorecards(
                pais=pais,
                criterio=criterio,
                sheet_names=sheet_names,
                pipeline_by_brand=pipeline_by_brand,
            )
            skipped_sheets.extend(calculation_skips)

            if not scorecards:
                print(f"{Colors.FAIL}No se pudo calcular ningún scorecard válido para '{source_file}'.{Colors.ENDC}")
                _print_skipped_sheets_summary(source_file, skipped_sheets)
                continue

            export_scorecards_single_sheet(scorecards, output_dir, output_name)
            _print_skipped_sheets_summary(source_file, skipped_sheets)
        except Exception as exc:
            print(f"{Colors.FAIL}Error procesando '{source_file}': {exc}{Colors.ENDC}")
            continue

    total_elapsed = time.perf_counter() - global_start
    print(f"\n{Colors.OKGREEN}Proceso finalizado en {total_elapsed:.1f}s.{Colors.ENDC}")


if __name__ == "__main__":
    main()
