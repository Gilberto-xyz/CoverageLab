import unittest
import os
from datetime import date

from openpyxl import load_workbook

import coverage_studio as studio


studio._load_heavy_modules()


def candidate(
    pipeline: int,
    correlation: float,
    variation_gap: float,
    *,
    trend_match: bool = True,
) -> studio.OptimalPipelineCandidate:
    return studio.OptimalPipelineCandidate(
        pipeline=pipeline,
        current_correlation=correlation,
        current_variation=0.10,
        wp_current_variation=0.10,
        variation_distance_points=variation_gap,
        current_trend_match=trend_match,
        previous_year_correlation=float("nan"),
        two_year_correlation=float("nan"),
        previous_year_variation=float("nan"),
        wp_previous_year_variation=float("nan"),
        historical_trend_match=False,
        recent_shipment_outlier=False,
        forced_by_sheet=False,
    )


class AutoPipelineComparisonTests(unittest.TestCase):
    def test_pipeline_report_row_preserves_resolved_sheet_metadata(self) -> None:
        bank_row = {
            "Fabricante/Marca": "Natura",
            "Cesta": "Cuidado Personal",
            "Codigo Categoria": "MAKE",
            "Categoria": "Maquillaje-Cosmeticos",
            "Periodo": date(2026, 5, 1),
            "Pipeline": 1,
        }
        variations = studio.pd.DataFrame(
            [
                {
                    "Tipo": "Anual",
                    **{f"Cliente P{pipeline}": 0.01 for pipeline in range(1, 7)},
                }
            ]
        )
        candidates = (candidate(1, 0.85, 0.2),)

        row = studio.build_pipeline_report_row(
            bank_row=bank_row,
            df_variations=variations,
            candidates=candidates,
            selection_reason="Pipeline indicado en el nombre de la hoja",
            ref_month_year="05-26",
        )

        self.assertEqual(row["Fabricante/Marca"], "Natura")
        self.assertEqual(row["Codigo Categoria"], "MAKE")
        self.assertEqual(row["Categoria"], "Maquillaje-Cosmeticos")
        self.assertEqual(row["Cesta"], "Cuidado Personal")
        self.assertEqual(row["Periodo"], date(2026, 5, 1))

    def test_pipeline_report_headers_are_localized_for_each_language(self) -> None:
        self.assertEqual(
            studio.localize_pipeline_report_header("Correlación P2", "ES"),
            "Correlación P2",
        )
        self.assertEqual(
            studio.localize_pipeline_report_header("Correlación P2", "PT"),
            "Correlação P2",
        )
        self.assertEqual(
            studio.localize_pipeline_report_header("Correlación P2", "EN"),
            "Correlation P2",
        )
        self.assertEqual(
            studio.localize_pipeline_report_header("Cobertura 05-26", "EN"),
            "Coverage 05-26",
        )
        self.assertEqual(
            studio.localize_pipeline_report_header("Codigo Categoria", "ES"),
            "Código de categoría",
        )
        self.assertEqual(
            studio.localize_pipeline_report_header("Categoria", "EN"),
            "Category",
        )

    def test_saved_pipeline_report_uses_percent_correlations_and_uniform_rows(self) -> None:
        columns = studio.build_pipeline_report_columns("05-26")
        row = {column: "" for column in columns}
        row.update(
            {
                "Fabricante/Marca": "Marca teste",
                "Cesta": "Cuidado Personal",
                "Codigo Categoria": "FRAG",
                "Categoria": "Fragancias",
                "Periodo": date(2026, 5, 1),
                "Pipeline": 2,
                "Raw Buyers Media Ano Mov Atual": 100,
                "% VAR WP by Numerator": 5.5,
                "Correlación P1": 0.965,
                "Correlación P2": 0.855,
                "Correlación seleccionada": 0.855,
                "Correlación top": 0.965,
                "Pipeline AUTO Correlación": 1,
                "Correlación AUTO Correlación": 0.965,
                "Pipeline AUTO Balanceado": 2,
                "Correlación AUTO Balanceado": 0.855,
                "Pérdida de correlación Balanceado": 0.110,
            }
        )
        negative_row = dict(row)
        negative_row["% VAR WP by Numerator"] = -3.2
        report_df = studio.pd.DataFrame([row, negative_row], columns=columns)

        temp_dir = os.path.join(
            os.path.dirname(studio.__file__),
            f".test_pipeline_report_{os.getpid()}",
        )
        os.makedirs(temp_dir, exist_ok=True)
        path = ""
        try:
            path = studio.save_pipeline_report(
                df_pipeline_report=report_df,
                carpeta_salida=temp_dir,
                fabricante="Marca",
                categoria_nombre="Categoria",
                categoria_nombre_corto="Categoria",
                pais_nombre="Brasil",
                ref_month_year="05-26",
                coverage_label="Cobertura Absoluta",
                language_code="PT",
            )
            workbook = load_workbook(path)
            worksheet = workbook["Relatório Pipelines"]
            headers = {
                str(cell.value): cell.column
                for cell in worksheet[1]
                if cell.value is not None
            }

            correlation_col = headers["Correlação P1"]
            numerator_col = headers["% VAR WP pela Numerator"]
            self.assertEqual(
                worksheet.cell(2, headers["Código da categoria"]).value,
                "FRAG",
            )
            self.assertEqual(
                worksheet.cell(2, headers["Categoria"]).value,
                "Fragancias",
            )
            self.assertEqual(
                worksheet.cell(2, headers["Cesta"]).value,
                "Cuidado Personal",
            )
            self.assertEqual(
                worksheet.cell(2, headers["Período"]).number_format,
                "mmm-yy",
            )
            self.assertTrue(
                worksheet.cell(1, numerator_col).fill.fgColor.rgb.endswith("2F75B5")
            )
            self.assertTrue(
                worksheet.cell(2, numerator_col).fill.fgColor.rgb.endswith("DDEBF7")
            )
            self.assertTrue(
                worksheet.cell(3, numerator_col).fill.fgColor.rgb.endswith("DDEBF7")
            )
            self.assertTrue(
                worksheet.cell(2, headers["Cesta"]).fill.fgColor.rgb.endswith("FFEBEB")
            )
            self.assertEqual(worksheet.cell(2, correlation_col).value, 0.965)
            self.assertEqual(worksheet.cell(2, correlation_col).number_format, "0.0%")
            self.assertEqual(
                worksheet.row_dimensions[1].height,
                studio.PIPELINE_REPORT_HEADER_ROW_HEIGHT,
            )
            self.assertEqual(
                worksheet.row_dimensions[2].height,
                studio.PIPELINE_REPORT_DATA_ROW_HEIGHT,
            )
            self.assertEqual(
                worksheet.row_dimensions[3].height,
                studio.PIPELINE_REPORT_DATA_ROW_HEIGHT,
            )
            workbook.close()
        finally:
            if path and os.path.isfile(path):
                os.remove(path)
            if os.path.isdir(temp_dir):
                os.rmdir(temp_dir)

    def test_material_variation_balance_prefers_shorter_p2_for_nosotras(self) -> None:
        candidates = (
            candidate(1, 0.869, 6.00),
            candidate(2, 0.855, 5.36),
            candidate(3, 0.886, 6.60),
            candidate(4, 0.916, 5.80),
            candidate(5, 0.965, 7.20),
            candidate(6, 0.937, 6.90),
        )

        chosen = studio._choose_material_variation_candidate(candidates)

        self.assertIsNotNone(chosen)
        self.assertEqual(chosen.pipeline, 2)
        reason = studio._material_variation_reason(chosen, candidates)
        self.assertIn("mejora gap=1.84pp", reason)
        self.assertEqual(
            studio._pipeline_decision_type(reason),
            "Balance material de variación",
        )
        diagnostics = studio.build_auto_pipeline_comparison_diagnostics(
            studio.AutoPipelineComparison(
                correlation=studio.select_correlation_pipeline(candidates),
                balanced=studio.OptimalPipelineSelection(
                    chosen.pipeline,
                    reason,
                    candidates,
                ),
            )
        )
        self.assertEqual(diagnostics["Pipeline AUTO Correlación"], 5)
        self.assertEqual(diagnostics["Pipeline AUTO Balanceado"], 2)
        self.assertEqual(diagnostics["Conflicto AUTO Correlación vs Balanceado"], "Medio")
        self.assertEqual(diagnostics["Mejora gap de variación Balanceado"], 1.84)

    def test_small_variation_improvement_does_not_override_correlation(self) -> None:
        candidates = (
            candidate(2, 0.676, 3.25),
            candidate(5, 0.750, 4.06),
        )

        self.assertIsNone(studio._choose_material_variation_candidate(candidates))

    def test_longer_pipeline_does_not_enter_material_balance(self) -> None:
        candidates = (
            candidate(2, 0.995, 2.04),
            candidate(6, 0.930, 0.78),
        )

        self.assertIsNone(studio._choose_material_variation_candidate(candidates))

    def test_auto_correlation_uses_the_highest_current_mat_correlation(self) -> None:
        candidates = (
            candidate(1, 0.80, 0.10),
            candidate(4, 0.40, 0.01),
            candidate(6, 0.92, 2.00, trend_match=False),
        )

        selection = studio.select_correlation_pipeline(candidates, forced_pipeline=1)

        self.assertEqual(selection.pipeline, 6)
        self.assertIn("máxima correlación MAT", selection.reason)

    def test_high_correlation_sacrifice_is_reported_for_balanced_override(self) -> None:
        candidates = (
            candidate(4, 0.387, 0.05),
            candidate(6, 0.925, 1.94),
        )
        comparison = studio.AutoPipelineComparison(
            correlation=studio.select_correlation_pipeline(candidates),
            balanced=studio.OptimalPipelineSelection(
                4,
                "ajuste casi exacto de variación anual",
                candidates,
            ),
        )

        diagnostics = studio.build_auto_pipeline_comparison_diagnostics(comparison)

        self.assertEqual(diagnostics["Pipeline AUTO Correlación"], 6)
        self.assertEqual(diagnostics["Pipeline AUTO Balanceado"], 4)
        self.assertEqual(diagnostics["Conflicto AUTO Correlación vs Balanceado"], "Alto")
        self.assertEqual(diagnostics["Revisión requerida"], "SI")
        self.assertEqual(
            diagnostics["Tipo de decisión AUTO Balanceado"],
            "Override por alineación de variación",
        )

    def test_same_pipeline_is_reported_without_conflict(self) -> None:
        candidates = (
            candidate(1, 0.70, 1.00),
            candidate(2, 0.90, 0.20),
        )
        correlation = studio.select_correlation_pipeline(candidates)
        comparison = studio.AutoPipelineComparison(
            correlation=correlation,
            balanced=studio.OptimalPipelineSelection(
                2,
                "correlación MAT de Año Actual positiva y variación anual alineada",
                candidates,
            ),
        )

        diagnostics = studio.build_auto_pipeline_comparison_diagnostics(comparison)

        self.assertEqual(diagnostics["Conflicto AUTO Correlación vs Balanceado"], "Sin conflicto")
        self.assertEqual(diagnostics["Revisión requerida"], "NO")

    def test_report_columns_include_both_auto_modes(self) -> None:
        columns = studio.build_pipeline_report_columns("05-26")

        self.assertEqual(
            columns[:5],
            ["Fabricante/Marca", "Cesta", "Codigo Categoria", "Categoria", "Periodo"],
        )
        self.assertIn("Pipeline AUTO Correlación", columns)
        self.assertIn("Pipeline AUTO Balanceado", columns)
        self.assertIn("Revisión requerida", columns)

    def test_report_columns_preserve_multiple_sheet_periods(self) -> None:
        columns = studio.build_pipeline_report_columns_for_periods(["05-26", "06-26"])

        self.assertIn("Cobertura 05-25", columns)
        self.assertIn("Cobertura 05-26", columns)
        self.assertIn("Cobertura 06-25", columns)
        self.assertIn("Cobertura 06-26", columns)
        self.assertEqual(len(columns), len(set(columns)))


if __name__ == "__main__":
    unittest.main()
