import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import coverage_studio as studio


class CoverageBankTablesTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        studio._load_heavy_modules()

    def _sample_bank(self) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {
                    "Periodo": pd.Timestamp("2026-06-01"),
                    "Fabricante": "Natura",
                    "Codigo Categoria": "FRAG",
                    "Categoria": "Fragancias",
                    "Fabricante/Marca": "Natura",
                    "Pipeline": 6,
                },
                {
                    "Periodo": pd.Timestamp("2026-06-01"),
                    "Fabricante": "Natura",
                    "Codigo Categoria": "MAKE",
                    "Categoria": "Maquillaje",
                    "Fabricante/Marca": "Natura",
                    "Pipeline": 6,
                },
            ]
        )

    def test_table_views_leave_two_blank_rows_between_headers(self) -> None:
        original, with_code, second_startrow = studio.build_coverage_bank_table_views(
            self._sample_bank()
        )

        self.assertNotIn("Codigo Categoria", original.columns)
        self.assertIn("Codigo Categoria", with_code.columns)
        self.assertEqual(second_startrow, len(original.index) + 3)

    def test_saved_bank_contains_both_tables_on_the_same_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = studio.save_coverage_bank(
                self._sample_bank(),
                carpeta_salida=temp_dir,
                nombre_base_archivo="Natura",
                fabricante="Natura",
                categoria_nombre="Cross Category",
                categoria_nombre_corto="Cross Category",
                pais_nombre="Mexico",
                ref_month_year="06-26",
                coverage_label="Cobertura",
                coverage_type="absoluta",
                coverage_slide_variant="classic",
            )

            self.assertTrue(Path(output_path).exists())
            workbook = load_workbook(output_path, data_only=False)
            worksheet = workbook["Sheet1"]

            top_headers = [cell.value for cell in worksheet[1] if cell.value is not None]
            bottom_header_row = len(self._sample_bank().index) + 4
            bottom_headers = [
                cell.value for cell in worksheet[bottom_header_row] if cell.value is not None
            ]

            self.assertNotIn("Codigo Categoria", top_headers)
            self.assertIn("Codigo Categoria", bottom_headers)
            self.assertTrue(
                all(
                    worksheet.cell(row=row, column=column).value is None
                    for row in (bottom_header_row - 2, bottom_header_row - 1)
                    for column in range(1, worksheet.max_column + 1)
                )
            )
            category_code_column = bottom_headers.index("Codigo Categoria") + 1
            self.assertEqual(
                worksheet.cell(
                    row=bottom_header_row + 1,
                    column=category_code_column,
                ).value,
                "FRAG",
            )


if __name__ == "__main__":
    unittest.main()
