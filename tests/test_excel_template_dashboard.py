import os
import unittest
from datetime import datetime

from openpyxl import Workbook, load_workbook

import coverage_studio as studio


studio._load_heavy_modules()


class ExcelTemplateDashboardTests(unittest.TestCase):
    def _test_directory(self, suffix: str) -> str:
        path = os.path.join(os.path.dirname(studio.__file__), f".test_excel_dashboard_{os.getpid()}_{suffix}")
        os.makedirs(path, exist_ok=True)
        return path

    def _remove_test_directory(self, path: str) -> None:
        for name in os.listdir(path):
            os.remove(os.path.join(path, name))
        os.rmdir(path)

    def _build_template(self, path: str) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "P4_Marca Demo"
        headers = [
            studio.COL_DATA,
            studio.COL_SELL_IN,
            studio.COL_SELL_OUT,
            studio.COL_COMPRA_MEDIA,
            studio.COL_COMPRA_OCA,
            studio.COL_FREQ,
            studio.COL_PENET,
            studio.COL_BUYERS,
            studio.COL_ANO,
            studio.COL_TRI,
            studio.COL_SEM,
            studio.COL_SELL_IN_SIM,
            studio.COL_ACUM_SELL_OUT,
            studio.COL_ACUM_SELL_IN,
            *[f"P{pipeline}" for pipeline in range(7)],
            studio.COL_EVO_KANTAR_YOY,
            studio.COL_EVO_SELLIN_YOY,
        ]
        sheet.append(headers)
        dates = studio.pd.date_range(datetime(2023, 1, 1), periods=40, freq="MS")
        for index, date_value in enumerate(dates, start=1):
            sell_in = 1000 + (index * 35) + (120 if index % 7 == 0 else 0)
            sell_out = 950 + (index * 30) + (90 if index % 9 == 0 else 0)
            sheet.append(
                [
                    date_value.to_pydatetime(),
                    sell_in,
                    sell_out,
                    4.5 + (index * 0.03),
                    2.2 + (index * 0.01),
                    2.0 + (index * 0.005),
                    20 + (index * 0.1),
                    1400 + (index * 8),
                    date_value.year,
                    ((date_value.month - 1) // 3) + 1,
                    ((date_value.month - 1) // 6) + 1,
                    sell_in,
                    sell_out * 12,
                    sell_in * 12,
                    *[75 + pipeline + (index * 0.05) for pipeline in range(7)],
                    0.01 * ((index % 5) - 2),
                    0.012 * ((index % 6) - 3),
                ]
            )
        workbook.save(path)

    def test_template_adds_pipeline_selector_kpi_charts_and_alerts(self) -> None:
        temp_dir = self._test_directory("features")
        try:
            path = os.path.join(temp_dir, "template.xlsx")
            self._build_template(path)

            studio.add_native_excel_charts(
                path,
                coverage_label="Cobertura Absoluta",
                trend_axis="simple",
                evolution_slide_variant="simple",
                include_english=False,
                pais_nombre="Mexico",
            )

            workbook = load_workbook(path, data_only=False)
            source = workbook["P4_Marca Demo"]

            self.assertFalse(any(sheet.title.startswith("Dash_") for sheet in workbook.worksheets))
            self.assertEqual(source["Z3"].value, 4)
            validations = list(source.data_validations.dataValidation)
            self.assertEqual(len(validations), 1)
            self.assertEqual(validations[0].formula1, '"0,1,2,3,4,5,6"')
            self.assertIsNone(source["AT1"].value)
            self.assertIsNone(source["AW1"].value)
            self.assertIsNone(source["BC1"].value)

            self.assertIn('="Sell-in (P"', source["X1"].value)
            self.assertIn("$Z$3", source["X1"].value)
            self.assertNotIn("INDEX", source["X2"].value)
            self.assertNotIn("ROW()", source["X2"].value)
            self.assertIn("IF($Z$3=0,$B$2", source["X2"].value)
            self.assertIn("$Z$3", source["X2"].value)
            self.assertFalse(any(name.startswith("TrendSellIn_") for name in workbook.defined_names))
            self.assertEqual(len(source._charts), 8)
            chart_sizes = {
                (chart.anchor.ext.width, chart.anchor.ext.height)
                for chart in source._charts
            }
            self.assertEqual(len(chart_sizes), 1)
            width_emu, height_emu = next(iter(chart_sizes))
            self.assertGreater(width_emu, 5328000)
            self.assertGreater(height_emu, 2592000)
            anchors = [
                (chart.anchor._from.col, chart.anchor._from.row)
                for chart in source._charts
            ]
            self.assertEqual(
                anchors,
                [(26, 4), (36, 4), (26, 24), (36, 24), (26, 44), (36, 44), (26, 64), (36, 64)],
            )
            for chart in source._charts:
                self.assertIsNotNone(chart.x_axis.title)
                self.assertIsNotNone(chart.y_axis.title)
                self.assertEqual(chart.x_axis.__class__.__name__, "DateAxis")
                self.assertEqual(chart.x_axis.tickLblPos, "low")
                self.assertFalse(chart.x_axis.delete)
                self.assertEqual(chart.y_axis.tickLblPos, "nextTo")
                self.assertFalse(chart.y_axis.delete)
                self.assertEqual(chart.x_axis.numFmt.formatCode, "mmm-yy")
                self.assertEqual(
                    chart.y_axis.majorGridlines.spPr.line.solidFill.srgbClr,
                    "E3E7ED",
                )

            coverage_chart = source._charts[0]
            self.assertEqual(len(coverage_chart._charts), 2)
            self.assertEqual(coverage_chart._charts[0].y_axis.axPos, "l")
            self.assertEqual(coverage_chart._charts[1].y_axis.axPos, "r")
            coverage_refs = [
                series.val.numRef.f
                for subchart in coverage_chart._charts
                for series in subchart.series
            ]
            coverage_categories = [
                series.cat.numRef.f
                for subchart in coverage_chart._charts
                for series in subchart.series
            ]
            self.assertEqual(
                coverage_refs,
                ["'P4_Marca Demo'!$S$2:$S$41", "'P4_Marca Demo'!$G$2:$G$41"],
            )
            self.assertEqual(
                coverage_categories,
                ["'P4_Marca Demo'!$A$2:$A$41", "'P4_Marca Demo'!$A$2:$A$41"],
            )

            trend_chart = source._charts[2]
            self.assertEqual(trend_chart.series[0].val.numRef.f, "'P4_Marca Demo'!$X$2:$X$41")
            self.assertEqual(trend_chart.series[1].val.numRef.f, "'P4_Marca Demo'!$C$2:$C$41")
            self.assertEqual(trend_chart.series[0].cat.numRef.f, "'P4_Marca Demo'!$A$2:$A$41")
            self.assertEqual(trend_chart.series[0].tx.v, "Sell-in (P0–P6)")

            expected_kpi_columns = ["D", "E", "F", "G", "H"]
            expected_kpi_formats = ["#,##0.0", "#,##0.00", "0.00", "0.0", "#,##0"]
            for chart, column, number_format in zip(
                source._charts[3:],
                expected_kpi_columns,
                expected_kpi_formats,
            ):
                self.assertEqual(chart.series[0].val.numRef.f, f"'P4_Marca Demo'!${column}$2:${column}$41")
                self.assertEqual(chart.series[0].cat.numRef.f, "'P4_Marca Demo'!$A$2:$A$41")
                self.assertFalse(chart.series[0].dLbls.showVal)
                self.assertEqual(chart.series[0].dLbls.dLblPos, "t")
                self.assertEqual(chart.series[0].dLbls.numFmt, number_format)
                self.assertEqual(
                    [label.idx for label in chart.series[0].dLbls.dLbl],
                    list(range(0, 40, 3)),
                )
                self.assertTrue(all(label.showVal for label in chart.series[0].dLbls.dLbl))
                self.assertEqual(chart.series[0].marker.symbol, "circle")
            self.assertEqual(
                [source.cell(row, 27).value for row in range(87, 92)],
                ["Compra media", "Compra por ocasión", "Frecuencia", "Penetración", "Buyers"],
            )
            self.assertIn("$AC$3", source["AF87"].value)
            self.assertIn("$AF$3", source["AF87"].value)
            self.assertEqual(workbook.active.title, source.title)
            workbook.close()
        finally:
            self._remove_test_directory(temp_dir)

    def test_pipeline_shift_formula_covers_p0_to_p6_without_invalid_rows(self) -> None:
        first_formula = studio._excel_pipeline_shift_formula(2)
        self.assertIn("IF($Z$3=0,$B$2", first_formula)
        self.assertNotIn("$B$1", first_formula)
        self.assertNotIn("INDEX", first_formula)
        self.assertNotIn("ROW()", first_formula)
        for pipeline in range(7):
            self.assertIn(f"$Z$3={pipeline}", first_formula)

        seventh_lag_formula = studio._excel_pipeline_shift_formula(8)
        for source_row in range(2, 9):
            self.assertIn(f"$B${source_row}", seventh_lag_formula)

    def test_pipeline_selector_is_generated_for_single_and_double_axis_modes(self) -> None:
        for axis_mode in ("simple", "doble"):
            with self.subTest(axis_mode=axis_mode):
                temp_dir = self._test_directory(f"axis_{axis_mode}")
                try:
                    path = os.path.join(temp_dir, "template.xlsx")
                    self._build_template(path)
                    studio.add_native_excel_charts(
                        path,
                        coverage_label="Cobertura Absoluta",
                        trend_axis=axis_mode,
                        evolution_slide_variant="simple",
                        include_english=False,
                        pais_nombre="Mexico",
                    )
                    workbook = load_workbook(path, data_only=False)
                    source = workbook["P4_Marca Demo"]
                    self.assertEqual(source["Z3"].value, 4)
                    self.assertIn("$Z$3=0", source["X8"].value)
                    self.assertIn("$Z$3=6", source["X8"].value)
                    self.assertEqual(
                        source._charts[2].series[0].val.numRef.f,
                        "'P4_Marca Demo'!$X$2:$X$41",
                    )
                    workbook.close()
                finally:
                    self._remove_test_directory(temp_dir)

    def test_autofit_supports_dashboard_merged_headers(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.merge_cells("A1:F1")
        sheet["A1"] = "Dashboard de diagnóstico"
        sheet["B2"] = "Desfase Sell-in (meses)"

        studio.autofit_worksheet_columns(sheet, min_width=10.0, max_width=30.0, padding=2.0)

        self.assertGreaterEqual(sheet.column_dimensions["A"].width, 10.0)
        self.assertGreaterEqual(sheet.column_dimensions["B"].width, 10.0)

    def test_template_dashboard_regeneration_is_idempotent(self) -> None:
        temp_dir = self._test_directory("idempotent")
        try:
            path = os.path.join(temp_dir, "template.xlsx")
            self._build_template(path)
            kwargs = {
                "coverage_label": "Cobertura Absoluta",
                "trend_axis": "simple",
                "evolution_slide_variant": "simple",
                "include_english": False,
                "pais_nombre": "Mexico",
            }

            studio.add_native_excel_charts(path, **kwargs)
            studio.add_native_excel_charts(path, **kwargs)

            workbook = load_workbook(path, data_only=False)
            self.assertFalse(any(sheet.title.startswith("Dash_") for sheet in workbook.worksheets))
            source = workbook["P4_Marca Demo"]
            self.assertEqual(len(source._charts), 8)
            self.assertEqual(len(source.data_validations.dataValidation), 1)
            self.assertFalse(any(name.startswith("TrendSellIn_") for name in workbook.defined_names))
            workbook.close()
        finally:
            self._remove_test_directory(temp_dir)


if __name__ == "__main__":
    unittest.main()
